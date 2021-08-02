import logging
import pathlib
import re
from dataclasses import dataclass
from typing import Union
from collections import defaultdict
from openpyxl import load_workbook

import system
import utils
from models import RunTarget

LOG = logging.getLogger('ledscreen.spreadsheets')
FUZZY_PATTERN = re.compile('[^a-z0-9]')


def fuzzy_text_compare(has: str, wants: str):
    if has is not None:
        has = has.lower()
        sanitized_has = FUZZY_PATTERN.sub('', has)
        wants = wants.lower()
        sanitized_wants = FUZZY_PATTERN.sub('', wants)
        return sanitized_wants in sanitized_has
    return False


def parse_cell_boolean(raw_value):
    if raw_value is None:
        return False

    if isinstance(raw_value, str):
        if fuzzy_text_compare(raw_value, 'true'):
            return True
        elif fuzzy_text_compare(raw_value, 'false'):
            return False
        elif fuzzy_text_compare(raw_value, 'yes'):
            return True
        elif fuzzy_text_compare(raw_value, 'no'):
            return False

    return raw_value


IMPORT_SHEET_NAME = 'Roster'
USERNAME_HEADER = 'Username'
PASSWORD_HEADER = 'Password'
EXPIRY_HEADER = 'Expiry'
LOCK_HEADER = 'Lock'
RUN_PRIVILEGE_HEADER = 'Run Privilege'
MAX_RUNTIME_HEADER = 'Max Runtime'
COMMENT_HEADER = 'Comment'

IMPORT_COLUMN_NAMES = [
    USERNAME_HEADER,
    PASSWORD_HEADER,
    EXPIRY_HEADER,
    LOCK_HEADER,
    RUN_PRIVILEGE_HEADER,
    MAX_RUNTIME_HEADER,
    COMMENT_HEADER
]

EXPORT_COLUMN_NAMES = [
    'ID',
    'Username',
    'Password',
    'Issue Date',
    'Expire Date',
    'Lock Date',
    'Last Login Date',
    'Last Logout Date',
    'Run Count',
    'Login Count',
    'Online Time',
    'Workspace ID',
    'Comment'
]


def load_student_roster(filename):
    path_obj = pathlib.Path(filename)
    book = load_workbook(filename)
    book_name = path_obj.name
    LOG.debug(f'loaded workbook "{book_name}"')

    values = defaultdict(list)

    if len(book.worksheets) > 0:
        sheet = book.worksheets[0]
        LOG.debug(f'"{book_name}": selected first sheet')

        for i, col in enumerate(sheet.iter_cols(values_only=True), start=1):
            header = col[0]
            rows = col[1:]
            values[header].extend(rows)
    else:
        LOG.debug(f'cannot import "{book_name}": no sheets')

    book.close()
    LOG.debug(f'closed workbook "{path_obj.name}"')

    return values


# from https://stackoverflow.com/questions/23861680/convert-spreadsheet-number-to-column-letter
def column_letters(n):
    letters = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


@dataclass(frozen=True)
class ValidationError:
    col: int
    row: int
    header: Union[None, str]
    message: str

    def cell_notation(self):
        return column_letters(self.col) + str(self.row)

    def __repr__(self):
        return f'<ValidationError at {self.cell_notation()} "{self.header}" "{self.message}">'


def validate_roster(roster_data: dict):
    errors = []

    column_count = len(IMPORT_COLUMN_NAMES)
    if len(roster_data.keys()) < column_count:
        errors.append(ValidationError(0, 0, None, f'worksheet must have {column_count} or more columns'))

    found_headers = []
    for ci, (header, values) in enumerate(roster_data.items(), start=1):
        if USERNAME_HEADER not in found_headers and fuzzy_text_compare(header, USERNAME_HEADER):
            for ri, value in enumerate(values, start=2):
                if value is not None:
                    if not system.validate_username(value):
                        errors.append(ValidationError(
                            ci,
                            ri,
                            USERNAME_HEADER,
                            'invalid username (1-40 characters, A-Z, a-z and spaces only)'))
                else:
                    errors.append(ValidationError(ci,
                                                  ri,
                                                  USERNAME_HEADER,
                                                  'username is required'))
            found_headers.append(USERNAME_HEADER)
        elif PASSWORD_HEADER not in found_headers and fuzzy_text_compare(header, PASSWORD_HEADER):
            for ri, value in enumerate(values, start=2):
                if value is not None:
                    if not system.validate_password(value):
                        errors.append(ValidationError(
                            ci,
                            ri,
                            PASSWORD_HEADER,
                            'invalid password (6-64 characters)'))
            found_headers.append(PASSWORD_HEADER)
        elif EXPIRY_HEADER not in found_headers and fuzzy_text_compare(header, EXPIRY_HEADER):
            for ri, value in enumerate(values, start=2):
                if value is not None:
                    if utils.parse_account_expiry(value) is None:
                        errors.append(ValidationError(
                            ci,
                            ri,
                            EXPIRY_HEADER,
                            f'invalid expiry date "{value}" (absolute future dates only)'))
            found_headers.append(EXPIRY_HEADER)
        elif RUN_PRIVILEGE_HEADER not in found_headers and fuzzy_text_compare(header, RUN_PRIVILEGE_HEADER):
            for ri, value in enumerate(values, start=2):
                if value is not None:
                    try:
                        RunTarget[value.upper()]
                    except KeyError:
                        errors.append(ValidationError(ci,
                                                      ri,
                                                      RUN_PRIVILEGE_HEADER,
                                                      f'invalid run privilege "{value}" (can be {", ".join(RunTarget.names())})'))
                else:
                    errors.append(ValidationError(ci,
                                                  ri,
                                                  RUN_PRIVILEGE_HEADER,
                                                  f'run privilege is required'))
            found_headers.append(RUN_PRIVILEGE_HEADER)
        elif MAX_RUNTIME_HEADER not in found_headers and fuzzy_text_compare(header, MAX_RUNTIME_HEADER):
            for ri, value in enumerate(values, start=2):
                if value is not None and value < 2:
                    errors.append(ValidationError(ci,
                                                  ri,
                                                  MAX_RUNTIME_HEADER,
                                                  f'max runtime must be at least 2 seconds or not defined'))
            found_headers.append(MAX_RUNTIME_HEADER)

        if len(found_headers) == len(IMPORT_COLUMN_NAMES):
            break
    return errors


def transform_roster(roster_data: dict):
    simplified = defaultdict(list)

    found_keys = []
    for ci, (header, values) in enumerate(roster_data.items(), start=1):
        if len(values) > 0:
            closest_key = None
            for column_name in IMPORT_COLUMN_NAMES:
                if fuzzy_text_compare(header, column_name):
                    LOG.debug(f'column {ci}: found {column_name} ({len(values)} data rows)')

                    if column_name in found_keys:
                        LOG.debug(f'column {ci}: duplicate known column name "{column_name}"')
                    else:
                        found_keys.append(column_name)
                        closest_key = column_name
                        break
            else:
                LOG.debug(f'column {ci}: ignoring column "{header}"')

            if closest_key is not None:
                if closest_key == EXPIRY_HEADER or closest_key == LOCK_HEADER or closest_key == RUN_PRIVILEGE_HEADER:
                    simplified_values = []
                    for value in values:
                        transformed_value = None

                        if closest_key == EXPIRY_HEADER:
                            transformed_value = utils.parse_account_expiry(value)
                        elif closest_key == LOCK_HEADER:
                            transformed_value = parse_cell_boolean(value)
                        elif closest_key == RUN_PRIVILEGE_HEADER:
                            transformed_value = RunTarget[value.upper()]

                        simplified_values.append(transformed_value)
                    simplified[closest_key].extend(simplified_values)
                else:
                    simplified[closest_key].extend(values)

    return simplified
